from openpyxl import Workbook 
from datetime import date

def patientReport(applist):
    wb = Workbook()
    wb.create_sheet("Sheet_zero")  

    ws1 = wb.active
    ws1['A1'] = 'ID'
    ws1['B1'] = 'FIRST NAME'
    ws1['C1'] = 'LAST NAME' 
    ws1['D1'] = 'EMAIL'
    ws1['E1'] = 'USERNAME'
    ws1['F1'] = 'BIRTHDATE'
 
    row_start = 2  
    col_start = 1   

    index = 0  

    for app in applist: 
        ws1.cell(row_start+index,col_start).value = app.user.id 
        ws1.cell(row_start+index,col_start+1).value = app.user.first_name 
        ws1.cell(row_start+index,col_start+2).value = app.user.last_name  
        ws1.cell(row_start+index,col_start+3).value = app.user.email
        ws1.cell(row_start+index,col_start+4).value = app.user.username  
        ws1.cell(row_start+index,col_start+5).value = app.user.profile.birth_date 
        index = index+1
    
    return wb 
