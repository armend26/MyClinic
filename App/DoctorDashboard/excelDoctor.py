from openpyxl import Workbook 
from datetime import date

def createReport(applist):
    wb = Workbook()
    wb.create_sheet("Sheet_zero")  

    ws1 = wb.active
    ws1['A1'] = 'ID'
    ws1['B1'] = 'DATE'
    ws1['C1'] = 'TIME' 
    ws1['D1'] = 'NOTE'

    row_start = 2  
    col_start = 1   

    index = 0  

    for app in applist: 
        if  str(app.date) < str(date.today()):
            continue
        else:
            ws1.cell(row_start+index,col_start).value = app.id 
            ws1.cell(row_start+index,col_start+1).value = app.date 
            ws1.cell(row_start+index,col_start+2).value = app.time   
            ws1.cell(row_start+index,col_start+3).value = app.note 
            index = index+1
    
    return wb 
