from openpyxl import Workbook 
from .timeslot import timeslot
 

def createReport(user,date,app):
    wb = Workbook()
    wb.create_sheet("Sheet_zero")  

  
    ws1 = wb.active
    ws1['B2'] = 'No' 
    ws1['C2'] = 'Time'
    ws1['D2'] = 'Status'  

    row_start = 3  #start below the header row 2
    col_start = 2  #starts from column B

    status_list_index = 0 

    for i in range(0,16):
        ws1.cell(row_start+i,col_start).value = i+1 
        ws1.cell(row_start+i,col_start+1).value = timeslot[i]
        try:
            if timeslot[i] in app[status_list_index]:
                ws1.cell(row_start+i,col_start+2).value = 'Reserved'
                app.append('')
                status_list_index = status_list_index+1
        except IndexError:
            pass
        

    return wb 

